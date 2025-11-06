# app/routers/export.py
from __future__ import annotations

from datetime import datetime, timedelta, date
import io
from typing import List, Dict, Optional, Tuple

from fastapi import APIRouter, Depends, HTTPException, Query, Response, Request
from starlette.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.routers.auth import require_roles
from app.db.session import get_db
from app.models.applicant import Applicant, ApplicantDoc
from app.models.checklist import ChecklistItem
from app.services.pdf_service import (
    render_single_pdf,
    render_single_pdf_a5,
    render_batch_pdf,
)

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.utils.soft_delete import exclude_deleted, ensure_not_deleted

# Audit
from app.services.audit import write_audit

router = APIRouter()  # khÃ´ng prefix; main sáº½ mount /api


# ------------------- Audit helper -------------------
def _audit_print_or_export(
    *,
    request: Request,
    db: Session,
    user,
    action: str,             # "PRINT" | "EXPORT"
    scope: str,              # "day" | "dot"
    filters: dict,
    count: int,
    status: str,             # "SUCCESS" | "FAIL"
    error: str | None = None,
    name_mode: str | None = None,   # "full"/"split"
    target_type: str = "ApplicantBatch",
    target_id: str | None = None
):
    payload = {
        "scope": scope,
        "filters": filters or {},
        "count": int(count or 0),
        "name_mode": name_mode,
        "actor_id": getattr(user, "id", None),
        "actor_name": getattr(user, "full_name", None) or getattr(user, "username", None),
    }
    if error:
        payload["error"] = str(error)

    try:
        write_audit(
            db,
            action=action,
            target_type=target_type,
            target_id=target_id,
            prev_values=None,
            new_values=payload,
            status=status,
            request=request,
        )
        db.commit()
    except Exception:
        try: db.rollback()
        except Exception: pass


# ================= Helpers chung =================
def _parse_day_any(raw: str) -> date:
    s = (raw or "").strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    raise HTTPException(
        status_code=400,
        detail="Sai Ä‘á»‹nh dáº¡ng ngÃ y. DÃ¹ng 'date=dd/MM/YYYY' (khuyáº¿n nghá»‹) hoáº·c 'day=YYYY-MM-DD'.",
    )

def _items_merged_by_versions(db: Session, version_ids: set) -> List[ChecklistItem]:
    code_seen = set()
    items: List[ChecklistItem] = []
    for vid in version_ids:
        q = db.query(ChecklistItem).filter(ChecklistItem.version_id == vid)
        if hasattr(ChecklistItem, "order_index"):
            q = q.order_by(getattr(ChecklistItem, "order_index").asc())
        elif hasattr(ChecklistItem, "order_no"):
            q = q.order_by(getattr(ChecklistItem, "order_no").asc())
        else:
            q = q.order_by(ChecklistItem.id.asc())
        for it in q.all():
            if it.code not in code_seen:
                code_seen.add(it.code)
                items.append(it)
    return items

def _docs_map_by_mssv(docs: List[ApplicantDoc]) -> Dict[str, Dict[str, int]]:
    out: Dict[str, Dict[str, int]] = {}
    for d in docs:
        out.setdefault(d.applicant_ma_so_hv, {})[d.code] = int(d.so_luong or 0)
    return out

def _fmt_date_excel(v: Optional[object]) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        v = v.date()
    if isinstance(v, date):
        return v.strftime("%d/%m/%Y")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%d/%m/%Y")
        except Exception:
            continue
    return s

def _display_name(a: Applicant) -> str:
    hd = (getattr(a, "ho_dem", None) or "").strip()
    t = (getattr(a, "ten", None) or "").strip()
    if hd or t:
        return f"{hd} {t}".strip()
    return (getattr(a, "ho_ten", None) or "").strip()

def _split_name(a: Applicant) -> Tuple[str, str]:
    hd = (getattr(a, "ho_dem", None) or "").strip()
    t  = (getattr(a, "ten", None) or "").strip()
    if hd or t:
        return hd, t
    full = (getattr(a, "ho_ten", None) or "").strip()
    if not full:
        return "", ""
    parts = full.split()
    if len(parts) == 1:
        return "", parts[0]
    return " ".join(parts[:-1]), parts[-1]


def _build_excel_bytes(
    apps: List[Applicant],
    docs: List[ApplicantDoc],
    items_all: List[ChecklistItem],
    *,
    split_name: bool = False,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ho so"

    base_headers = ["STT", "MÃ£ há»“ sÆ¡", "NgÃ y nháº­n", "Email há»c viÃªn"]
    if split_name:
        # giá»¯ "Há» vÃ  tÃªn" + thÃªm "Há» Ä‘á»‡m", "TÃªn"
        base_headers += ["Há» vÃ  tÃªn", "Há» Ä‘á»‡m", "TÃªn"]
    else:
        base_headers += ["Há» vÃ  tÃªn"]
    base_headers += [
        "MSHV", "NgÃ y sinh", "Sá»‘ ÄT", "NgÃ nh nháº­p há»c", "Äá»£t", "KhÃ³a",
        "ÄÃ£ TN trÆ°á»›c Ä‘Ã³", "Ghi chÃº", "NgÆ°á»i nháº­n (kÃ½ tÃªn)", "DÃ¢n tá»™c",
    ]

    item_headers = [getattr(it, "display_name", None) or it.code for it in items_all]
    headers = base_headers + item_headers
    ws.append(headers)

    docs_by_mssv = _docs_map_by_mssv(docs)

    for idx, a in enumerate(apps, start=1):
        common_prefix = [
            idx,
            a.ma_ho_so or "",
            _fmt_date_excel(getattr(a, "ngay_nhan_hs", None)),
            a.email_hoc_vien or "",
        ]

        if split_name:
            full = _display_name(a)
            ln, fn = _split_name(a)
            name_cells = [full, ln, fn]
        else:
            name_cells = [_display_name(a)]

        common_suffix = [
            a.ma_so_hv or "",
            _fmt_date_excel(getattr(a, "ngay_sinh", None)),
            a.so_dt or "",
            getattr(a, "nganh_nhap_hoc", None) or getattr(a, "nganh", None) or "",
            a.dot or "",
            getattr(a, "khoa", "") or "",
            a.da_tn_truoc_do or "",
            a.ghi_chu or "",
            a.nguoi_nhan_ky_ten or "",
            getattr(a, "dan_toc", None) or "",
        ]

        dm = docs_by_mssv.get(a.ma_so_hv, {})
        doc_row = [int(dm.get(it.code, 0)) for it in items_all]

        ws.append(common_prefix + name_cells + common_suffix + doc_row)

    ws.freeze_panes = "A2"
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _get_app_by_mssv(db: Session, ma_so_hv: str) -> Applicant:
    a = db.query(Applicant).filter(Applicant.ma_so_hv == ma_so_hv).first()
    if not a:
        raise HTTPException(status_code=404, detail="Applicant not found")
    ensure_not_deleted(a)  # ðŸ”’ há»“ sÆ¡ xoÃ¡ -> 410
    return a

def _get_items_for_app(db: Session, app: Applicant):
    ver_id = getattr(app, "checklist_version_id", None)
    q = db.query(ChecklistItem)
    if ver_id:
        q = q.filter(ChecklistItem.version_id == ver_id)
    if hasattr(ChecklistItem, "order_index"):
        q = q.order_by(getattr(ChecklistItem, "order_index").asc())
    elif hasattr(ChecklistItem, "order_no"):
        q = q.order_by(getattr(ChecklistItem, "order_no").asc())
    else:
        q = q.order_by(ChecklistItem.id.asc())
    return q.all()

def _get_docs_for_mssv(db: Session, ma_so_hv: str):
    return db.query(ApplicantDoc).filter(ApplicantDoc.applicant_ma_so_hv == ma_so_hv).all()


# ================= EXPORT EXCEL THEO NGÃ€Y =================
@router.get("/export/excel")
def export_excel(
    request: Request,
    day: str | None = Query(None, description="YYYY-MM-DD"),
    date_q: str | None = Query(None, alias="date", description="dd/MM/YYYY"),
    name: str = Query("split", description="Kiá»ƒu cá»™t tÃªn: 'full' hoáº·c 'split'"),
    db: Session = Depends(get_db),
    user=Depends(require_roles("Admin", "NhanVien")),
):
    raw = date_q or day
    if not raw:
        _audit_print_or_export(
            request=request, db=db, user=user,
            action="EXPORT", scope="day",
            filters={"date": None}, count=0,
            status="FAIL", error="MISSING_DATE", name_mode=name
        )
        raise HTTPException(status_code=400, detail="Thiáº¿u tham sá»‘ 'date=dd/MM/YYYY' hoáº·c 'day=YYYY-MM-DD'")
    d = _parse_day_any(raw)

    d1 = datetime.combine(d, datetime.min.time())
    d2 = d1 + timedelta(days=1)

    q = db.query(Applicant).filter(Applicant.ngay_nhan_hs >= d1, Applicant.ngay_nhan_hs < d2)
    q = exclude_deleted(Applicant, q)
    apps = q.order_by(Applicant.created_at.asc(), Applicant.ma_so_hv.asc()).all()

    if not apps:
        q = exclude_deleted(Applicant, db.query(Applicant).filter(Applicant.ngay_nhan_hs == d))
        apps = q.order_by(Applicant.created_at.asc(), Applicant.ma_so_hv.asc()).all()

    apps = [a for a in apps if ensure_not_deleted(a, raise_http_exception=False)]
    if not apps:
        _audit_print_or_export(
            request=request, db=db, user=user,
            action="EXPORT", scope="day",
            filters={"date": d.isoformat()}, count=0,
            status="FAIL", error="NO_DATA", name_mode=name
        )
        raise HTTPException(status_code=404, detail=f"KhÃ´ng cÃ³ há»“ sÆ¡ trong ngÃ y {d.strftime('%d/%m/%Y')}")

    mssv_list = [a.ma_so_hv for a in apps]
    docs = db.query(ApplicantDoc).filter(ApplicantDoc.applicant_ma_so_hv.in_(mssv_list)).all()

    version_ids = {a.checklist_version_id for a in apps if a.checklist_version_id}
    items_all = _items_merged_by_versions(db, version_ids) if version_ids else []

    split = (name == "split")
    xls_bytes = _build_excel_bytes(apps, docs, items_all, split_name=split)
    suffix = "_split" if split else ""
    filename = f"Export_{d.strftime('%d-%m-%Y')}{suffix}.xlsx"

    # Audit OK
    _audit_print_or_export(
        request=request, db=db, user=user,
        action="EXPORT", scope="day",
        filters={"date": d.isoformat()}, count=len(apps),
        status="SUCCESS", name_mode=("split" if split else "full")
    )

    return StreamingResponse(
        io.BytesIO(xls_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


# ================= EXPORT EXCEL THEO Äá»¢T =================
@router.get("/export/excel-dot")
def export_excel_dot(
    request: Request,
    dot: str = Query(..., description="VÃ­ dá»¥: 'Äá»£t 1/2025' hoáº·c '9'"),
    khoa: str | None = Query(None, description="(Tuá»³ chá»n) Lá»c theo KhÃ³a, vÃ­ dá»¥: '27'"),
    name: str = Query("split", description="Kiá»ƒu cá»™t tÃªn: 'full' hoáº·c 'split'"),
    db: Session = Depends(get_db),
    user=Depends(require_roles("Admin", "NhanVien")),
):
    key = (dot or "").strip()
    if not key:
        _audit_print_or_export(
            request=request, db=db, user=user,
            action="EXPORT", scope="dot",
            filters={"dot": None, "khoa": (khoa or "")}, count=0,
            status="FAIL", error="MISSING_DOT", name_mode=name
        )
        raise HTTPException(status_code=400, detail="Thiáº¿u tham sá»‘ 'dot'")

    q = (
        db.query(Applicant)
        .filter(Applicant.dot.isnot(None))
        .filter(Applicant.dot.ilike(f"%{key}%"))
    )
    if (khoa or "").strip():
        k = khoa.strip()
        q = q.filter(Applicant.khoa.isnot(None)).filter(func.lower(func.trim(Applicant.khoa)) == k.lower())

    q = exclude_deleted(Applicant, q)

    apps = q.order_by(Applicant.created_at.asc(), Applicant.ma_so_hv.asc()).all()
    apps = [a for a in apps if ensure_not_deleted(a, raise_http_exception=False)]
    if not apps:
        _audit_print_or_export(
            request=request, db=db, user=user,
            action="EXPORT", scope="dot",
            filters={"dot": key, "khoa": (khoa or "")}, count=0,
            status="FAIL", error="NO_DATA", name_mode=name
        )
        raise HTTPException(status_code=404, detail="KhÃ´ng cÃ³ há»“ sÆ¡ nÃ o phÃ¹ há»£p")

    mssv_list = [a.ma_so_hv for a in apps]
    docs = db.query(ApplicantDoc).filter(ApplicantDoc.applicant_ma_so_hv.in_(mssv_list)).all()

    items_all = _items_merged_by_versions(db, {a.checklist_version_id for a in apps if a.checklist_version_id})

    split = (name == "split")
    xls_bytes = _build_excel_bytes(apps, docs, items_all, split_name=split)
    safe_dot = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in key)
    safe_khoa = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in (khoa or ""))
    suffix = f"{safe_dot}" + (f"_Khoa_{safe_khoa}" if safe_khoa else "")
    suffix2 = "_split" if split else ""
    filename = f"Export_Dot_{suffix}{suffix2}.xlsx"

    # Audit OK
    _audit_print_or_export(
        request=request, db=db, user=user,
        action="EXPORT", scope="dot",
        filters={"dot": key, "khoa": (khoa or "")}, count=len(apps),
        status="SUCCESS", name_mode=("split" if split else "full")
    )

    return StreamingResponse(
        io.BytesIO(xls_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


# ================= PRINT 1 Há»’ SÆ  (theo MSSV) =================
@router.get("/print/a5/{ma_so_hv}", summary="In 01 há»“ sÆ¡ A5 (ngang) theo MSSV")
def print_a5(ma_so_hv: str, db: Session = Depends(get_db)):
    app = _get_app_by_mssv(db, ma_so_hv)  # Ä‘Ã£ cháº·n deleted
    items = _get_items_for_app(db, app)
    docs = _get_docs_for_mssv(db, ma_so_hv)
    pdf_bytes = render_single_pdf_a5(app, items, docs)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename=\"{app.ma_ho_so or ma_so_hv}_A5.pdf\"'},
    )

@router.get("/print/a4/{ma_so_hv}", summary="In 01 há»“ sÆ¡ A4 (dá»c) theo MSSV")
def print_a4(ma_so_hv: str, db: Session = Depends(get_db)):
    app = _get_app_by_mssv(db, ma_so_hv)
    items = _get_items_for_app(db, app)
    docs = _get_docs_for_mssv(db, ma_so_hv)
    pdf_bytes = render_single_pdf(app, items, docs)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename=\"{app.ma_ho_so or ma_so_hv}_A4.pdf\"'},
    )
