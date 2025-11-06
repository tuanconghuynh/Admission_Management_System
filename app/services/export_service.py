# ================================
# app/services/export_service.py
# ================================
from __future__ import annotations
from typing import List, Dict, Iterable, Any, Optional, Tuple, Union, Callable
from io import BytesIO
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from ..models import Applicant, ApplicantDoc, ChecklistItem

DOC_PREFIX = "doc_"

# ---------- Helpers ----------
def _parse_to_date(v: Optional[object]) -> Optional[date]:
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

def _norm_gender(v: Optional[object]) -> str:
    """
    Chuẩn hoá giới tính về: 'Nam' | 'Nữ' | 'Khác' | ''.
    Hỗ trợ các biến thể: M/F, male/female, 1/0, nam/nu, ...
    """
    if v in (None, ""):
        return ""
    s = str(v).strip().lower()
    if s in {"1", "m", "male", "nam"}:
        return "Nam"
    if s in {"0", "f", "female", "nu", "nữ", "nư"}:
        return "Nữ"
    if s in {"other", "khac", "khác"}:
        return "Khác"
    if s == "nam":
        return "Nam"
    if s in {"nu", "nữ"}:
        return "Nữ"
    return s.capitalize()  # fallback

def _autosize(ws):
    ws.freeze_panes = "A2"
    for col in ws.columns:
        w = max(10, *(len(str(c.value)) if c.value else 0 for c in col)) + 2
        ws.column_dimensions[col[0].column_letter].width = min(w, 40)

def _get(obj: Union[dict, Any], key: str, default=None):
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)

def _display_name_from_obj(obj: Union[dict, Any]) -> str:
    """
    Ưu tiên ho_dem + ten; fallback ho_ten.
    Hỗ trợ cả dict (từ query thô) lẫn ORM object.
    """
    hd = (_get(obj, "ho_dem") or "").strip() if _get(obj, "ho_dem") is not None else ""
    t = (_get(obj, "ten") or "").strip() if _get(obj, "ten") is not None else ""
    if hd or t:
        return f"{hd} {t}".strip()
    return (_get(obj, "ho_ten") or "").strip()

def _split_name_cells(obj: Union[dict, Any]) -> Tuple[str, str]:
    """
    Trả về (ho_dem, ten) (có thể rỗng). Nếu chỉ có ho_ten mà không có tách,
    sẽ cố gắng tách nhẹ theo khoảng trắng cuối (không quá gắt).
    """
    ln = (_get(obj, "ho_dem") or "").strip() if _get(obj, "ho_dem") is not None else ""
    fn = (_get(obj, "ten") or "").strip() if _get(obj, "ten") is not None else ""
    if ln or fn:
        return ln, fn
    # Fallback tách từ ho_ten (best-effort)
    full = (_get(obj, "ho_ten") or "").strip()
    if not full:
        return "", ""
    parts = full.split()
    if len(parts) == 1:
        return "", parts[0]
    return " ".join(parts[:-1]), parts[-1]

def _set_date_format_by_header(ws, headers: List[str], header_names: Iterable[str]):
    """
    Đặt number_format = dd/mm/yyyy + căn giữa cho các cột ngày
    dựa trên tên header (an toàn khi thêm/bớt cột).
    """
    name_to_idx = {h: i + 1 for i, h in enumerate(headers)}  # 1-based
    for hn in header_names:
        col = name_to_idx.get(hn)
        if not col:
            continue
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
            for c in cell:
                if isinstance(c.value, (date, datetime)):
                    c.number_format = "dd/mm/yyyy"
                    c.alignment = Alignment(horizontal="center")

# ---------- Export 1: có cột checklist ----------
def build_excel_bytes_by_items(
    apps: List[Applicant],
    docs: List[ApplicantDoc],
    items: List[ChecklistItem],
    *,
    split_name: bool = False,
) -> bytes:
    """
    Xuất bảng có cột checklist.
    - split_name=False: 1 cột 'Họ tên'
    - split_name=True : 2 cột 'Họ và tên đệm' + 'Tên'
    """
    docs_by_mssv: Dict[str, Dict[str, int]] = {}
    for d in docs:
        docs_by_mssv.setdefault(d.applicant_ma_so_hv, {})[d.code] = int(d.so_luong or 0)

    base_headers = [
        "Ngày nhận HS", "Niên Khóa", "Mã hồ sơ", "Mã số HV",
    ]
    if split_name:
        base_headers += ["Họ và tên đệm", "Tên"]
    else:
        base_headers += ["Họ tên"]

    base_headers += [
        "Giới tính", "Dân tộc",
        "Email học viên", "Ngày sinh", "Số ĐT", "Ngành nhập học",
        "Đợt", "Đối tượng", "Ghi chú", "Printed"
    ]
    doc_headers = [f"{DOC_PREFIX}{it.code}" for it in (items or [])]
    headers = base_headers + doc_headers

    wb = Workbook()
    ws = wb.active
    ws.title = "Data_TongNgay"
    ws.append(headers)

    for a in apps:
        dm = docs_by_mssv.get(a.ma_so_hv, {})

        prefix = [
            _parse_to_date(a.ngay_nhan_hs),
            getattr(a, "khoa", ""),
            a.ma_ho_so or "",
            a.ma_so_hv or "",
        ]

        if split_name:
            ln, fn = _split_name_cells(a)
            name_cells = [ln, fn]
        else:
            name_cells = [_display_name_from_obj(a)]

        suffix = [
            _norm_gender(getattr(a, "gioi_tinh", "")),
            getattr(a, "dan_toc", "") or "",
            getattr(a, "email_hoc_vien", "") or "",
            _parse_to_date(getattr(a, "ngay_sinh", None)),
            a.so_dt or "",
            getattr(a, "nganh_nhap_hoc", None) or getattr(a, "nganh", None) or "",
            a.dot or "",
            a.da_tn_truoc_do or "",
            a.ghi_chu or "",
            bool(a.printed),
        ]

        row = prefix + name_cells + suffix
        for it in items or []:
            qty = int(dm.get(it.code, 0))
            row.append("" if qty == 0 else qty)
        ws.append(row)

    # format cột ngày theo tên header
    _set_date_format_by_header(ws, headers, header_names=["Ngày nhận HS", "Ngày sinh"])

    _autosize(ws)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ---------- Export 2: bảng đơn giản ----------
def build_excel_bytes_simple(
    rows: Iterable[Any],
    *,
    split_name: bool = False,
) -> bytes:
    """
    Xuất bảng tổng hợp đơn giản.
    - split_name=False: 1 cột 'Họ tên'
    - split_name=True : 2 cột 'Họ và tên đệm' + 'Tên'
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "TongHop"

    headers = [
        "Mã HS",
    ]
    if split_name:
        headers += ["Họ và tên đệm", "Tên"]
    else:
        headers += ["Họ tên"]

    headers += [
        "MSHV", "Giới tính", "Dân tộc",
        "Email học viên",
        "Ngày nhận HS", "Ngày sinh", "Ngành", "Đợt",
        "Khóa", "Người nhận", "Ghi chú"
    ]
    ws.append(headers)

    for a in rows:
        get = a.get if isinstance(a, dict) else lambda k, d=None: getattr(a, k, d)

        prefix = [get("ma_ho_so")]
        if split_name:
            ln, fn = _split_name_cells(a)
            name_cells = [ln, fn]
        else:
            name_cells = [_display_name_from_obj(a)]

        suffix = [
            get("ma_so_hv"),
            _norm_gender(get("gioi_tinh", "")),
            get("dan_toc", "") or "",
            get("email_hoc_vien", ""),
            _parse_to_date(get("ngay_nhan_hs")),
            _parse_to_date(get("ngay_sinh")),
            get("nganh_nhap_hoc") or get("nganh"),
            get("dot"),
            get("khoa"),
            get("nguoi_nhan_ky_ten"),
            get("ghi_chu"),
        ]

        ws.append(prefix + name_cells + suffix)

    # format cột ngày theo tên header
    _set_date_format_by_header(ws, headers, header_names=["Ngày nhận HS", "Ngày sinh"])

    _autosize(ws)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
